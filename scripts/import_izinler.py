#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
İK MUTTAŞ — Yıllık İzin Import Scripti
=========================================
6 Excel dosyasından personel yıllık izin verilerini sisteme aktarır.

Tek seferlik kullanım için yazılmıştır (idempotent DEĞİL):
- izin_gecmisi tablosunu TRUNCATE eder
- personel tablosuna hakkedis_yillara_gore JSONB kolonu ekler (yoksa)
- Her personel için yıllık hakediş + izin kullanım kayıtlarını import eder
- toplam_izin_hak ve kalan_izin alanlarını günceller
- Eşleşmeyen personelleri CSV olarak raporlar (import edilmez)

Kullanım:
    python import_izinler.py                  # DRY RUN (raporla, DB'ye dokunma)
    python import_izinler.py --apply          # Gerçekten uygula
"""
import os
import re
import sys
import json
import argparse
import asyncio
import asyncpg
import openpyxl
from datetime import datetime, date
from collections import Counter

EXCEL_DIR = "imports/yillik-izinler-2026-05"
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://muttas:CHANGE@127.0.0.1:5432/muttas_db")

SKIP_SHEET_NAMES = {"SAYFA11", "SAYFA12", "XXX"}
SKIP_SHEET_PATTERNS = ["İŞTEN ÇIK", "DEVRED", "DEVİR", "TÜM PERSONEL"]
GARBAGE_NAME_KEYWORDS = ['İSTİFA', 'İSTIFA', 'GÜN ÜCRETSİZ', 'RAPORLU', 'TOPLAM ',
                        'YILLIK ÜCRETLİ', 'İZİN CETVELİ', 'EYT', 'KOVULDU',
                        'ÇIKIŞ', 'AYRILDI', '...']

def normalize_isim(ad):
    """İsim eşleştirme için normalize: parantez içi sil, trim, çift boşluk, upper.
    
    DİKKAT: Parantez içi silinince 'EMRE KAYA (TC:42226826170)' → 'EMRE KAYA'
    olur. Aynı isimde farklı kişiler birbirine karışabilir. İmport sırasında
    bu durum tespit edilirse uyarı verilir.
    """
    if not ad: return ''
    s = str(ad)
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'\s+', ' ', s).strip().upper()
    return s

def is_garbage_name(ad):
    if not ad: return True
    s = ad.upper()
    if re.search(r'\d{2}[./,]\d{2}[./,]\d{4}', s): return True
    if re.search(r'\d{4}[-/]\d{2}[-/]\d{2}', s): return True
    for kw in GARBAGE_NAME_KEYWORDS:
        if kw in s: return True
    if len(s.split()) < 2: return True
    return False

def should_skip_sheet(sheet_name, all_sheets):
    name_u = sheet_name.strip().upper()
    if name_u in SKIP_SHEET_NAMES: return True
    for pat in SKIP_SHEET_PATTERNS:
        if pat in name_u: return True
    if "GÜNCEL" not in name_u:
        for other in all_sheets:
            if other.strip().upper() == name_u + " GÜNCEL":
                return True
    return False

def find_column_layout(ws, baslik_row):
    """Bir sheet'in sütun düzenini çıkar.
    
    Returns: {
        'izne_baslama': col_idx,
        'goreve_donus': col_idx,
        'gun_sayisi': col_idx,
        'bakiye': col_idx (opsiyonel),
        'yil_kolonlari': {col_idx: 2017, ...}
    }
    """
    layout = {'izne_baslama': None, 'goreve_donus': None, 'gun_sayisi': None,
              'bakiye': None, 'yil_kolonlari': {}}
    
    for c in range(1, ws.max_column + 1):
        r1 = ws.cell(row=baslik_row, column=c).value
        r2 = ws.cell(row=baslik_row + 1, column=c).value
        r1_str = str(r1).strip().upper() if r1 else ''
        r2_str = str(r2).strip().upper() if r2 else ''
        # Normalize boşlukları (multi-line başlıklar var: 'İzin\n Bitiş\n Tarihi')
        r2_norm = re.sub(r'\s+', ' ', r2_str)
        
        if 'İZNE BAŞLAMA' in r2_norm:
            layout['izne_baslama'] = c
        elif 'GÖREVE BAŞLAMA' in r2_norm or 'GÖREVE DÖNÜŞ' in r2_norm or 'İZİN BİTİŞ' in r2_norm:
            layout['goreve_donus'] = c
        elif 'GÜN SAYISI' in r2_norm:
            layout['gun_sayisi'] = c
        elif 'BAKİYE' in r1_str or 'KALAN İZİN' in r1_str:
            layout['bakiye'] = c
        
        # Yıl sütunları (R(baslik_row+1)'de yıl rakamı: 2015-2030)
        try:
            yil = int(r2) if r2 else None
            if yil and 2010 <= yil <= 2035:
                layout['yil_kolonlari'][c] = yil
        except (ValueError, TypeError):
            pass
    
    return layout

def parse_excel_files(excel_dir):
    """Tüm Excel'leri oku, personel kayıtları döndür."""
    sonuc = []
    sheet_uyari = []
    
    for fname in sorted(os.listdir(excel_dir)):
        if not fname.endswith('.xlsx'): continue
        wb = openpyxl.load_workbook(os.path.join(excel_dir, fname), data_only=True)
        
        for sheet_name in wb.sheetnames:
            if should_skip_sheet(sheet_name, wb.sheetnames): continue
            ws = wb[sheet_name]
            
            # Adı Soyadı başlık satırını bul
            baslik_row = None
            for r in range(1, min(20, ws.max_row + 1)):
                v = ws.cell(row=r, column=2).value
                if v and isinstance(v, str) and v.strip() == "Adı Soyadı":
                    baslik_row = r
                    break
            if baslik_row is None: continue
            
            data_start = baslik_row + 2
            layout = find_column_layout(ws, baslik_row)
            
            if not layout['izne_baslama']:
                sheet_uyari.append(f"{fname[:25]}/{sheet_name}: 'İzne Başlama' sütunu bulunamadı, kullanımlar okunamayacak")
                # Yine de personel listesini al
            
            current = None
            for r in range(data_start, ws.max_row + 1):
                ad = ws.cell(row=r, column=2).value
                unvan = ws.cell(row=r, column=3).value
                hak_tarih = ws.cell(row=r, column=4).value
                
                # Yeni personel başlık satırı
                if ad and isinstance(ad, str) and ad.strip():
                    ad_str = ad.strip()
                    if is_garbage_name(ad_str):
                        current = None
                        continue
                    
                    # Yıllık hakedişler
                    yillik = {}
                    for col_idx, yil in layout['yil_kolonlari'].items():
                        val = ws.cell(row=r, column=col_idx).value
                        if val is not None:
                            try:
                                gun = int(val)
                                if gun > 0:
                                    yillik[yil] = gun
                            except (ValueError, TypeError):
                                pass
                    
                    current = {
                        'isim_orijinal': ad_str,
                        'isim_norm': normalize_isim(ad_str),
                        'unvan': str(unvan).strip() if unvan else None,
                        'hakedis_tarihi': hak_tarih if isinstance(hak_tarih, (date, datetime)) else None,
                        'yillik_hakedisler': yillik,
                        'toplam_hakedis': sum(yillik.values()),
                        'kullanimlar': [],
                        'kaynak_dosya': fname,
                        'kaynak_sheet': sheet_name,
                    }
                    sonuc.append(current)
                
                # Kullanım satırı (current personel altında)
                if current and layout['izne_baslama']:
                    izne_bas = ws.cell(row=r, column=layout['izne_baslama']).value
                    goreve_don = ws.cell(row=r, column=layout['goreve_donus']).value if layout['goreve_donus'] else None
                    gun_val = ws.cell(row=r, column=layout['gun_sayisi']).value if layout['gun_sayisi'] else None
                    
                    if isinstance(izne_bas, (date, datetime)):
                        # Gün sayısı: önce X kolonu, yoksa W-V farkı
                        gun = None
                        if isinstance(gun_val, (int, float)) and gun_val > 0:
                            gun = int(gun_val)
                        elif isinstance(goreve_don, (date, datetime)):
                            d1 = izne_bas.date() if isinstance(izne_bas, datetime) else izne_bas
                            d2 = goreve_don.date() if isinstance(goreve_don, datetime) else goreve_don
                            gun = (d2 - d1).days
                        
                        if gun and gun > 0:
                            current['kullanimlar'].append({
                                'baslama': izne_bas.date() if isinstance(izne_bas, datetime) else izne_bas,
                                'donus': (goreve_don.date() if isinstance(goreve_don, datetime) else goreve_don) if isinstance(goreve_don, (date, datetime)) else None,
                                'gun': gun,
                            })
    
    if sheet_uyari:
        print("⚠️  Sheet uyarıları:")
        for u in sheet_uyari[:10]:
            print(f"   - {u}")
    
    return sonuc

async def main(dry_run):
    print(f"🔄 Excel dosyaları okunuyor ({EXCEL_DIR})...")
    if not os.path.isdir(EXCEL_DIR):
        print(f"❌ HATA: Excel dizini bulunamadı: {EXCEL_DIR}")
        return 1
    
    excel_data = parse_excel_files(EXCEL_DIR)
    print(f"   ✓ {len(excel_data)} personel kaydı okundu")
    toplam_kullanim = sum(len(p['kullanimlar']) for p in excel_data)
    print(f"   ✓ Toplam {toplam_kullanim} izin kullanım kaydı")
    
    # Aynı normalize isim, farklı orijinal isim çakışması (parantez içi TC vs)
    norm_to_orijinal = {}
    for p in excel_data:
        if p['isim_norm'] not in norm_to_orijinal:
            norm_to_orijinal[p['isim_norm']] = set()
        norm_to_orijinal[p['isim_norm']].add(p['isim_orijinal'])
    cakismalar = {k: v for k, v in norm_to_orijinal.items() if len(v) > 1}
    if cakismalar:
        print(f"\n⚠️  PARANTEZ TEMİZLİĞİ NEDENİYLE ÇAKIŞMA — Aynı normalize ad altında farklı orijinal adlar:")
        for norm, orijinaller in list(cakismalar.items())[:8]:
            print(f"   '{norm}' ← {sorted(orijinaller)}")
        print(f"   ... toplam {len(cakismalar)} çakışma")
        print(f"   Bu personeller AYNI DB kaydına bağlanır — eğer farklı kişilerse veri karışır.")
    
    conn = await asyncpg.connect(DATABASE_URL)
    
    try:
        # DB personel listesi
        rows = await conn.fetch("SELECT id, ad_soyad, tc_kimlik FROM personel WHERE aktif = TRUE")
        db_isim_index = {}
        for row in rows:
            norm = normalize_isim(row['ad_soyad'])
            if norm in db_isim_index:
                print(f"   ⚠️  DB'de mükerrer normalize isim: '{norm}' (id {db_isim_index[norm]['id']} ve {row['id']})")
            db_isim_index[norm] = dict(row)
        print(f"\n   ✓ DB'de {len(db_isim_index)} tekil normalize aktif isim ({len(rows)} kayıttan)")
        
        eski_kayit = await conn.fetchval("SELECT COUNT(*) FROM izin_gecmisi")
        print(f"   ✓ izin_gecmisi'nde mevcut {eski_kayit} kayıt (TRUNCATE edilecek)")
        
        col_exists = await conn.fetchval("""
            SELECT 1 FROM information_schema.columns
            WHERE table_name='personel' AND column_name='hakkedis_yillara_gore'
        """)
        print(f"   ✓ hakkedis_yillara_gore kolonu: {'VAR' if col_exists else 'YOK (eklenecek)'}")
        
        # izin_gecmisi şeması
        izin_cols = await conn.fetch("""
            SELECT column_name, data_type FROM information_schema.columns
            WHERE table_name = 'izin_gecmisi' ORDER BY ordinal_position
        """)
        izin_col_names = [r['column_name'] for r in izin_cols]
        print(f"   ✓ izin_gecmisi kolonları: {izin_col_names}")
        
        # Şema doğrulama
        if 'personel_id' not in izin_col_names:
            print(f"❌ HATA: izin_gecmisi'nde personel_id kolonu yok!")
            return 1
        
        # Eşleştirme
        eslesti = []
        eslesmedi = []
        for p in excel_data:
            if p['isim_norm'] in db_isim_index:
                eslesti.append((p, db_isim_index[p['isim_norm']]))
            else:
                eslesmedi.append(p)
        
        print(f"\n📊 Eşleştirme sonucu:")
        print(f"   ✅ Eşleşti:    {len(eslesti)} personel")
        print(f"   ❌ Eşleşmedi:  {len(eslesmedi)} personel (atlanacak)")
        eslesti_kullanim = sum(len(p['kullanimlar']) for p, _ in eslesti)
        eslesmedi_kullanim = sum(len(p['kullanimlar']) for p in eslesmedi)
        print(f"      → import edilecek izin: {eslesti_kullanim}")
        print(f"      → atlanan izin:        {eslesmedi_kullanim}")
        
        # Aynı DB kaydına eşleşen birden fazla Excel kaydı?
        db_id_count = Counter(db_p['id'] for _, db_p in eslesti)
        coklu_eslesme = {db_id: c for db_id, c in db_id_count.items() if c > 1}
        if coklu_eslesme:
            print(f"\n⚠️  Aynı DB personeline birden fazla Excel kaydı eşleşmiş ({len(coklu_eslesme)} kişi):")
            for db_id, count in list(coklu_eslesme.items())[:5]:
                excel_kayitlar = [p for p, dbp in eslesti if dbp['id'] == db_id]
                print(f"   DB ID {db_id} ({excel_kayitlar[0]['isim_orijinal']}): {count} Excel kaydı")
                for ek in excel_kayitlar:
                    print(f"      → {ek['kaynak_sheet']} ({len(ek['kullanimlar'])} kullanım)")
            print(f"   Bu durumda hakedişler MERGE edilir, kullanımlar BİRLEŞTİRİLİR.")
        
        # Rapor: eşleşmeyenler
        rapor_yolu = "import_eslesmeyenler.csv"
        with open(rapor_yolu, "w", encoding="utf-8") as f:
            f.write("isim,unvan,kaynak_dosya,kaynak_sheet,toplam_hakedis,kullanim_sayisi\n")
            for p in eslesmedi:
                f.write(f'"{p["isim_orijinal"]}","{p["unvan"] or ""}","{p["kaynak_dosya"]}",'
                       f'"{p["kaynak_sheet"]}",{p["toplam_hakedis"]},{len(p["kullanimlar"])}\n')
        print(f"\n   📄 Eşleşmeyenler raporu yazıldı: {rapor_yolu}")
        
        if dry_run:
            print(f"\n🟡 DRY RUN — DB'ye dokunulmadı.")
            print(f"   Gerçekten uygulamak için: python import_izinler.py --apply")
            return 0
        
        # === GERÇEK IMPORT ===
        print(f"\n🚀 IMPORT BAŞLIYOR")
        async with conn.transaction():
            if not col_exists:
                await conn.execute("ALTER TABLE personel ADD COLUMN hakkedis_yillara_gore JSONB")
                print(f"   ✓ personel.hakkedis_yillara_gore kolonu eklendi")
            
            await conn.execute("TRUNCATE TABLE izin_gecmisi RESTART IDENTITY")
            print(f"   ✓ izin_gecmisi temizlendi ({eski_kayit} kayıt silindi)")
            
            # Aynı DB ID'ye birden fazla Excel kaydı → birleştir
            from collections import defaultdict
            by_db_id = defaultdict(list)
            for excel_p, db_p in eslesti:
                by_db_id[db_p['id']].append(excel_p)
            
            personel_updated = 0
            insert_count = 0
            
            # izin_gecmisi için INSERT — kolon adlarını otomatik tespit et
            if 'baslangic' in izin_col_names and 'bitis' in izin_col_names:
                col_bas, col_bit = 'baslangic', 'bitis'
            elif 'baslangic_tarihi' in izin_col_names and 'bitis_tarihi' in izin_col_names:
                col_bas, col_bit = 'baslangic_tarihi', 'bitis_tarihi'
            elif 'baslama_tarihi' in izin_col_names and 'donus_tarihi' in izin_col_names:
                col_bas, col_bit = 'baslama_tarihi', 'donus_tarihi'
            else:
                raise RuntimeError(f"izin_gecmisi tablo şeması beklenenden farklı: {izin_col_names}")
            print(f"   ✓ izin_gecmisi INSERT kolonları: {col_bas}, {col_bit}")

            for db_id, excel_kayitlar in by_db_id.items():
                # Tüm hakedişleri birleştir (yıl bazında topla)
                merged_yillik = {}
                for ek in excel_kayitlar:
                    for yil, gun in ek['yillik_hakedisler'].items():
                        merged_yillik[yil] = merged_yillik.get(yil, 0) + gun

                # Tüm kullanımları birleştir
                tum_kullanim = []
                for ek in excel_kayitlar:
                    tum_kullanim.extend(ek['kullanimlar'])

                toplam_hak = sum(merged_yillik.values())
                toplam_kul = sum(k['gun'] for k in tum_kullanim)
                kalan = toplam_hak - toplam_kul

                # personel update
                await conn.execute("""
                    UPDATE personel SET
                        hakkedis_yillara_gore = $1::jsonb,
                        toplam_izin_hak = $2,
                        kalan_izin = $3
                    WHERE id = $4
                """, json.dumps(merged_yillik), toplam_hak, kalan, db_id)
                personel_updated += 1

                # izin_gecmisi insertions
                for k in tum_kullanim:
                    bitis_val = k['donus']
                    if bitis_val is None and k['baslama'] is not None:
                        from datetime import timedelta
                        bitis_val = k['baslama'] + timedelta(days=k['gun'])
                    await conn.execute(
                        f"INSERT INTO izin_gecmisi (personel_id, {col_bas}, {col_bit}, gun_sayisi) VALUES ($1, $2, $3, $4)",
                        db_id, k['baslama'], bitis_val, k['gun'])
                    insert_count += 1
            
            print(f"   ✓ {personel_updated} personel kaydı güncellendi")
            print(f"   ✓ {insert_count} izin_gecmisi kaydı eklendi")
        
        # Son doğrulama (transaction dışı)
        final_count = await conn.fetchval("SELECT COUNT(*) FROM izin_gecmisi")
        final_personel = await conn.fetchval(
            "SELECT COUNT(*) FROM personel WHERE hakkedis_yillara_gore IS NOT NULL")
        print(f"\n✅ IMPORT TAMAMLANDI")
        print(f"   izin_gecmisi: {final_count} kayıt")
        print(f"   personel.hakkedis_yillara_gore dolu: {final_personel} kayıt")
    finally:
        await conn.close()
    return 0

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Gerçekten uygula (yoksa dry-run)")
    parser.add_argument("--excel-dir", default=None, help="Excel dosyaları dizini")
    args = parser.parse_args()
    if args.excel_dir:
        EXCEL_DIR = args.excel_dir
    sys.exit(asyncio.run(main(dry_run=not args.apply)))
